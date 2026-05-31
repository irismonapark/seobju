import math
import calendar
import datetime
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter
import holidays as holidays_lib

from utils.errors import UserFacingError


HOURLY_WAGE = 10320
NIGHT_START = 22 * 60  # 1320분

# ── 컬러 상수 ──────────────────────────────────────────
SAT_FILL   = PatternFill('solid', fgColor='DDEEFF')
SUN_FILL   = PatternFill('solid', fgColor='FFDDDD')
SAT_FONT   = Font(color='0000CC', name='맑은 고딕', size=9)
SUN_FONT   = Font(color='CC0000', name='맑은 고딕', size=9)
NORM_FONT  = Font(name='맑은 고딕', size=9)
BOLD_FONT  = Font(name='맑은 고딕', size=9, bold=True)

HEADER_FILL = PatternFill("solid", fgColor="1A237E")
HEADER_FONT = Font(color="FFFFFF", bold=True, name='맑은 고딕', size=9)

ROW_FILLS = [
    PatternFill('solid', fgColor='FFFFFF'),
    PatternFill('solid', fgColor='F5F5F5'),
    PatternFill('solid', fgColor='EBEBEB'),
    PatternFill('solid', fgColor='F5F5F5'),
    PatternFill('solid', fgColor='EBEBEB'),
    PatternFill('solid', fgColor='E8E8E8'),
]

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

# 직원형태 정렬 순서 (A=사대, D=일용, F=프리)
EMP_TYPE_ORDER = {'A': 0, 'D': 1, 'F': 2}


def _thin():
    return Side(style='thin', color='AAAAAA')


def _medium():
    return Side(style='medium', color='444444')


def _make_border(left=True, right=True, top=True, bottom=True, thick_bottom=False):
    bot = _medium() if thick_bottom else (_thin() if bottom else Side())
    return Border(
        left=_thin()   if left   else Side(),
        right=_thin()  if right  else Side(),
        top=_thin()    if top    else Side(),
        bottom=bot,
    )


def _parse_time(time_str):
    """시간 문자열을 (분, 자정넘김여부)로 변환. '+HH:MM'은 자정 넘김."""
    import re
    if not time_str:
        return None, False
    time_str = str(time_str).strip()
    if not time_str or time_str in ('nan', 'NaN', 'None', ''):
        return None, False
    next_day = time_str.startswith('+')
    time_str = time_str.lstrip('+')
    m = re.match(r'(\d{1,2})[:\-](\d{2})', time_str)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2)), next_day
    m2 = re.match(r'(\d{3,4})$', time_str)
    if m2:
        t = time_str.zfill(4)
        return int(t[:2]) * 60 + int(t[2:]), next_day
    return None, False


def calc_hours(in_str, out_str, work_type):
    """출퇴근 시간으로 근무 시간 계산 (점심 1시간 자동 공제)."""
    zero = {'basic': 0, 'extend': 0, 'special': 0, 'special_ot': 0,
            'night_1_5': 0, 'night_2_0': 0}

    if not in_str or not out_str:
        return zero

    start_min, _       = _parse_time(in_str)
    end_min, is_next_day = _parse_time(out_str)

    if start_min is None or end_min is None:
        return zero

    if is_next_day or end_min <= start_min:
        end_min += 24 * 60

    total_min = end_min - start_min - 60
    if total_min <= 0:
        return zero

    is_holiday = work_type not in ('평일', '')

    night_1_5 = 0
    night_2_0 = 0
    basic_end_min = start_min + 8 * 60

    if basic_end_min > NIGHT_START:
        n1_start = max(start_min, NIGHT_START)
        n1_end   = min(end_min, basic_end_min)
        night_1_5 = max(0, n1_end - n1_start) // 60

    if end_min > NIGHT_START:
        if basic_end_min <= NIGHT_START:
            n2_start = max(start_min + 8 * 60, NIGHT_START)
            n2_end   = end_min
            night_2_0 = max(0, n2_end - n2_start) // 60
        else:
            n2_start = basic_end_min
            n2_end   = end_min
            if n2_end > NIGHT_START and n2_start < n2_end:
                night_2_0 = max(0, n2_end - max(n2_start, NIGHT_START)) // 60

    if is_holiday:
        basic_work  = min(total_min, 8 * 60)
        extend_work = max(0, total_min - 8 * 60)
        return {
            'basic': 0,
            'extend': 0,
            'special': basic_work // 60,
            'special_ot': extend_work // 60,
            'night_1_5': night_1_5,
            'night_2_0': night_2_0,
        }
    else:
        basic_work  = min(total_min, 8 * 60)
        extend_work = max(0, total_min - 8 * 60)
        return {
            'basic': basic_work // 60,
            'extend': extend_work // 60,
            'special': 0,
            'special_ot': 0,
            'night_1_5': night_1_5,
            'night_2_0': night_2_0,
        }


def get_weekly_holiday_bonus(year, month, employee_records):
    """주휴수당 대상 일요일 반환."""
    result = set()
    _, days_in_month = calendar.monthrange(year, month)
    first_day = datetime.date(year, month, 1)
    last_day  = datetime.date(year, month, days_in_month)

    kr_holidays = holidays_lib.country_holidays('KR', years=year)

    current = first_day
    while current.weekday() != 0:
        current += datetime.timedelta(days=1)

    while current <= last_day:
        sunday = current + datetime.timedelta(days=6)

        week_weekdays = [
            current + datetime.timedelta(days=i)
            for i in range(5)
            if (current + datetime.timedelta(days=i)).month == month
        ]

        if sunday.month == month and week_weekdays:
            all_worked = True
            for d in week_weekdays:
                rec   = employee_records.get(d.day, {})
                dtype = rec.get('근무일명칭', '')
                if d in kr_holidays:
                    continue
                if dtype and dtype not in ('평일', '휴일'):
                    continue
                in_  = rec.get('출근', '')
                out_ = rec.get('퇴근', '')
                if str(in_) in ('', 'nan', 'None') and str(out_) in ('', 'nan', 'None'):
                    all_worked = False
                    break
            if all_worked:
                result.add(sunday.day)

        current += datetime.timedelta(days=7)

    return result


def _get_holiday_names(year, month):
    kr_holidays = holidays_lib.country_holidays('KR', years=year)
    result = {}
    _, days_in_month = calendar.monthrange(year, month)
    for d in range(1, days_in_month + 1):
        dt = datetime.date(year, month, d)
        if dt in kr_holidays:
            result[d] = kr_holidays[dt]
    return result


def _get_day_type(year, month, day):
    """'weekday' / 'saturday' / 'sunday_holiday' 반환"""
    d = datetime.date(year, month, day)
    kr_holidays = holidays_lib.country_holidays('KR', years=year)
    if d in kr_holidays or d.weekday() == 6:
        return 'sunday_holiday'
    if d.weekday() == 5:
        return 'saturday'
    return 'weekday'


def _calc_employee_totals(emp, year, month, days_in_month, holiday_map):
    """직원별 월간 합계 계산."""
    days = emp['days']

    weekly_bonus_days = get_weekly_holiday_bonus(year, month, days)

    kr_holidays = holidays_lib.country_holidays('KR', years=year)

    totals = {k: 0 for k in ('basic', 'extend', 'special', 'special_ot',
                              'night_1_5', 'night_2_0')}
    day_results = {}

    for d in range(1, days_in_month + 1):
        if d in weekly_bonus_days:
            result = {'basic': 8, 'extend': 0, 'special': 0,
                      'special_ot': 0, 'night_1_5': 0, 'night_2_0': 0}
            day_results[d] = result
            totals['basic'] += 8
            continue

        day_info  = days.get(d, {})
        in_str    = day_info.get('출근', '')
        out_str   = day_info.get('퇴근', '')
        work_type = day_info.get('근무일명칭', '')

        date_obj  = datetime.date(year, month, d)
        is_pkg_legal_hol  = (date_obj in kr_holidays)
        is_name_legal_hol = (bool(work_type) and work_type not in ('평일', '휴일'))
        is_legal_hol      = is_pkg_legal_hol or is_name_legal_hol

        if is_pkg_legal_hol and not work_type:
            work_type = holiday_map.get(d, kr_holidays[date_obj])
        elif is_pkg_legal_hol:
            work_type = holiday_map.get(d, kr_holidays[date_obj])

        no_attend = (not in_str  or str(in_str).strip()  in ('', 'nan', 'None')) and \
                    (not out_str or str(out_str).strip() in ('', 'nan', 'None'))

        if is_legal_hol and no_attend:
            result = {'basic': 8, 'extend': 0, 'special': 0,
                      'special_ot': 0, 'night_1_5': 0, 'night_2_0': 0}
        else:
            result = calc_hours(in_str, out_str, work_type)

        day_results[d] = result
        for k in ('basic', 'extend', 'special', 'special_ot', 'night_1_5', 'night_2_0'):
            totals[k] += result[k]

    return totals, day_results, weekly_bonus_days


def generate_bc_excel(parsed_data, output_path, company=''):
    year      = parsed_data['year']
    month     = parsed_data['month']
    employees = list(parsed_data['employees'])

    if not employees:
        raise UserFacingError("변환할 직원 데이터가 없습니다.")

    # 직원형태별 정렬: A(사대) → D(일용) → F(프리) → 기타
    employees = sorted(employees, key=lambda e: EMP_TYPE_ORDER.get(e.get('emp_type', ''), 3))

    _, days_in_month = calendar.monthrange(year, month)

    company_suffix = f'_{company}' if company else ''
    b_tab_name = f"{year}-{month:02d} 근태{company_suffix}"
    c_tab_name = f"{year}-{month:02d} 청구내역서{company_suffix}"

    wb    = openpyxl.Workbook()
    ws_b  = wb.active
    ws_b.title = b_tab_name
    ws_c  = wb.create_sheet(title=c_tab_name)

    holiday_map = _get_holiday_names(year, month)

    emp_totals = []
    for emp in employees:
        totals, day_results, wb_days = _calc_employee_totals(
            emp, year, month, days_in_month, holiday_map)
        emp_totals.append((totals, day_results, wb_days))

    _build_b_tab(ws_b, employees, year, month, days_in_month, holiday_map, emp_totals)
    _build_c_tab(ws_c, employees, year, month, b_tab_name, emp_totals)

    wb.save(output_path)


# ─────────────────────────────────────────────────────────
#  B 탭
# ─────────────────────────────────────────────────────────
DOW_KR = ['월', '화', '수', '목', '금', '토', '일']

HDR_DAY_ROW  = 5
HDR_DOW_ROW  = 6
DATA_ROW_0   = 7

COL_SEQ   = 1
COL_NAME  = 2
COL_WAGE  = 3
COL_BLANK = 4   # 직원형태 표시용
COL_LABEL = 5
COL_DAY1  = 6
COL_SOGEUP = COL_DAY1 + 31
COL_SUM    = COL_DAY1 + 32


def _apply_day_style(cell, day_type, row_fill=None):
    if day_type == 'saturday':
        cell.fill = SAT_FILL
        if cell.font and cell.font.bold:
            cell.font = Font(color='0000CC', name='맑은 고딕', size=9, bold=True)
        else:
            cell.font = SAT_FONT
    elif day_type == 'sunday_holiday':
        cell.fill = SUN_FILL
        if cell.font and cell.font.bold:
            cell.font = Font(color='CC0000', name='맑은 고딕', size=9, bold=True)
        else:
            cell.font = SUN_FONT
    else:
        if row_fill:
            cell.fill = row_fill


def _build_b_tab(ws, employees, year, month, days_in_month, holiday_map, emp_totals,
                 _kr_hol_cache={}):
    _key = (year, month)
    if _key not in _kr_hol_cache:
        _kr_hol_cache[_key] = holidays_lib.country_holidays('KR', years=year)
    kr_holidays_obj = _kr_hol_cache[_key]

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 5   # 직원형태
    ws.column_dimensions['E'].width = 13
    for ci in range(COL_DAY1, COL_DAY1 + 32):
        ws.column_dimensions[get_column_letter(ci)].width = 4
    ws.column_dimensions[get_column_letter(COL_SOGEUP)].width = 5
    ws.column_dimensions[get_column_letter(COL_SUM)].width = 8

    title_cell = ws.cell(row=1, column=1, value=f"{year}년 {month:02d}월 근태현황")
    title_cell.font = Font(name='맑은 고딕', size=12, bold=True, color='1A237E')
    title_cell.alignment = CENTER
    last_col = COL_SUM
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col)

    for c, label in [(COL_SEQ,'순번'),(COL_NAME,'이름'),(COL_WAGE,'시급'),
                     (COL_BLANK,'유형'),(COL_LABEL,'구분'),
                     (COL_SOGEUP,'소급'),(COL_SUM,'합계')]:
        cell = ws.cell(row=3, column=c, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = _make_border()
        ws.merge_cells(start_row=3, start_column=c, end_row=6, end_column=c)

    day_group_cell = ws.cell(row=3, column=COL_DAY1, value='일  자  별  근  무  시  간')
    day_group_cell.font      = HEADER_FONT
    day_group_cell.fill      = HEADER_FILL
    day_group_cell.alignment = CENTER
    day_group_cell.border    = _make_border()
    ws.merge_cells(start_row=3, start_column=COL_DAY1,
                   end_row=4, end_column=COL_DAY1 + 30)

    for r in [3, 4]:
        ws.row_dimensions[r].height = 14

    ws.row_dimensions[HDR_DAY_ROW].height = 14
    for d in range(1, days_in_month + 1):
        col   = COL_DAY1 + d - 1
        dtype = _get_day_type(year, month, d)
        cell  = ws.cell(row=HDR_DAY_ROW, column=col, value=d)
        cell.font   = HEADER_FONT if dtype == 'weekday' else (
            Font(color='0000CC', name='맑은 고딕', size=9, bold=True) if dtype=='saturday'
            else Font(color='CC0000', name='맑은 고딕', size=9, bold=True))
        cell.fill   = HEADER_FILL if dtype == 'weekday' else (
            SAT_FILL if dtype=='saturday' else SUN_FILL)
        cell.alignment = CENTER
        cell.border    = _make_border()

    for d in range(days_in_month + 1, 32):
        col = COL_DAY1 + d - 1
        ws.cell(row=HDR_DAY_ROW, column=col).border = _make_border()

    ws.row_dimensions[HDR_DOW_ROW].height = 14
    for d in range(1, days_in_month + 1):
        col   = COL_DAY1 + d - 1
        dtype = _get_day_type(year, month, d)
        dow   = DOW_KR[datetime.date(year, month, d).weekday()]
        cell  = ws.cell(row=HDR_DOW_ROW, column=col, value=dow)
        cell.alignment = CENTER
        cell.border    = _make_border()
        _apply_day_style(cell, dtype)
        if dtype == 'weekday':
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    for d in range(days_in_month + 1, 32):
        ws.cell(row=HDR_DOW_ROW, column=COL_DAY1+d-1).border = _make_border()

    for emp_idx, (emp, (totals, day_results, wb_days)) in enumerate(zip(employees, emp_totals)):
        name     = emp['name']
        emp_type = emp.get('emp_type', '')
        days     = emp['days']
        base_row = DATA_ROW_0 + emp_idx * 6
        is_last  = (emp_idx == len(employees) - 1)

        night_h = totals['night_1_5'] + totals['night_2_0']

        row_labels = ['기본급', '연장', '주특', '특잔수당', '심야', '근태']
        row_keys   = ['basic', 'extend', 'special', 'special_ot', 'night', 'attendance']

        for r_off, (label, key) in enumerate(zip(row_labels, row_keys)):
            r         = base_row + r_off
            thick_bot = (r_off == 5)
            row_fill  = ROW_FILLS[r_off]
            ws.row_dimensions[r].height = 15

            for col, val in [(COL_SEQ,   emp_idx+1 if r_off==0 else ''),
                             (COL_NAME,  name if r_off==0 else ''),
                             (COL_WAGE,  HOURLY_WAGE if r_off==0 else ''),
                             (COL_BLANK, emp_type if r_off==0 else ''),
                             (COL_LABEL, label)]:
                cell = ws.cell(r, col, val)
                cell.font      = NORM_FONT
                cell.alignment = RIGHT if col==COL_WAGE else CENTER
                cell.fill      = row_fill
                cell.border    = _make_border(thick_bottom=thick_bot)
                if col == COL_WAGE and r_off == 0:
                    cell.number_format = '#,##0'

            sogeup_cell = ws.cell(r, COL_SOGEUP)
            sogeup_cell.fill      = row_fill
            sogeup_cell.border    = _make_border(thick_bottom=thick_bot)
            sogeup_cell.alignment = CENTER

            sum_cell = ws.cell(r, COL_SUM)
            sum_cell.font      = NORM_FONT
            sum_cell.fill      = row_fill
            sum_cell.border    = _make_border(thick_bottom=thick_bot)
            sum_cell.alignment = CENTER
            if key == 'attendance':
                sum_cell.value = '계'
            elif key == 'night':
                nt = totals['night_1_5'] + totals['night_2_0']
                sum_cell.value = nt if nt > 0 else 0
            else:
                sum_cell.value = totals.get(key, 0)

            for d in range(1, 32):
                col  = COL_DAY1 + d - 1
                cell = ws.cell(r, col)
                cell.font      = NORM_FONT
                cell.alignment = CENTER
                cell.border    = _make_border(thick_bottom=thick_bot)

                if d <= days_in_month:
                    dtype = _get_day_type(year, month, d)
                    _apply_day_style(cell, dtype, row_fill)

                    if key == 'attendance':
                        근태_text = None
                        # A파일 근태 컬럼값 우선
                        a_attendance = emp.get('attendance', {})
                        if d in a_attendance:
                            근태_text = a_attendance[d]
                        else:
                            dt = datetime.date(year, month, d)
                            if dt in kr_holidays_obj:
                                근태_text = kr_holidays_obj[dt]
                        cell.value = 근태_text
                    elif key == 'night':
                        val = day_results.get(d, {})
                        nt  = val.get('night_1_5', 0) + val.get('night_2_0', 0)
                        cell.value = nt if nt > 0 else None
                    else:
                        val = day_results.get(d, {}).get(key, 0)
                        cell.value = val if val > 0 else None
                else:
                    cell.fill = PatternFill('solid', fgColor='DDDDDD')


# ─────────────────────────────────────────────────────────
#  C 탭
#
#  컬럼 레이아웃 (1-indexed):
#   B(2)=순번  C(3)=성명  D(4)=직원형태  E(5)=시급
#   F(6)=기본급시간  G(7)=기본급금액
#   H(8)=연장시간    I(9)=연장금액
#   J(10)=주특시간   K(11)=주특금액
#   L(12)=특잔시간   M(13)=특잔금액
#   N(14)=심야시간   O(15)=심야금액
#   P(16)=간식대     Q(17)=상여금
#   R(18)=직접비소계 S(19)=일반관리비 T(20)=기업이윤
#   U(21)=간접비소계 V(22)=총액
# ─────────────────────────────────────────────────────────
def _build_c_tab(ws_c, employees, year, month, b_tab_name, emp_totals):
    BODY_FONT = Font(name='맑은 고딕', size=9)
    BOLD9     = Font(name='맑은 고딕', size=9, bold=True)
    MONEY_FMT = '#,##0'

    ws_c.column_dimensions['B'].width = 5
    ws_c.column_dimensions['C'].width = 10
    ws_c.column_dimensions['D'].width = 7   # 직원형태
    ws_c.column_dimensions['E'].width = 9   # 시급
    for l in ['F','H','I','J','K','L','M','N','O','P','Q']:
        ws_c.column_dimensions[l].width = 9
    for l in ['G','R','S','T','U','V']:
        ws_c.column_dimensions[l].width = 14

    ws_c['B1'] = f"{year}년 {month}월 청구내역서"
    ws_c['B1'].font = Font(name='맑은 고딕', bold=True, size=12)
    ws_c['S1'] = 0.03   # 일반관리비율
    ws_c['T1'] = 0.03   # 기업이윤율
    ws_c['V1'] = "사업장:"

    row3_headers = {
        'B':'순번','C':'성명','D':'직원형태','E':'시급',
        'F':'기본급','H':'연장','J':'주특',
        'L':'특잔수당','N':'심야','P':'간식대',
        'Q':'상여금','R':'직접비소계','S':'일반관리비',
        'T':'기업이윤','U':'간접비소계','V':'총액'
    }
    row4_headers = {
        'F':'시간','G':'금액','H':'시간','I':'금액',
        'J':'시간','K':'금액','L':'시간','M':'금액',
        'N':'시간','O':'금액'
    }

    for col_letter, val in row3_headers.items():
        cell = ws_c[f"{col_letter}3"]
        cell.value     = val
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = _make_border()

    for col_letter, val in row4_headers.items():
        cell = ws_c[f"{col_letter}4"]
        cell.value     = val
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = _make_border()

    for emp_idx, (emp, (totals, day_results, wb_days)) in enumerate(zip(employees, emp_totals)):
        name     = emp['name']
        emp_type = emp.get('emp_type', '')

        night_amount  = int(totals['night_1_5'] * HOURLY_WAGE * 1.5 +
                            totals['night_2_0'] * HOURLY_WAGE * 2.0)

        basic_h       = totals['basic']
        extend_h      = totals['extend']
        special_h     = totals['special']
        special_ot_h  = totals['special_ot']
        night_h       = totals['night_1_5'] + totals['night_2_0']

        basic_a      = basic_h * HOURLY_WAGE
        extend_a     = int(extend_h * HOURLY_WAGE * 1.5)
        special_a    = int(special_h * HOURLY_WAGE * 1.5)
        special_ot_a = int(special_ot_h * HOURLY_WAGE * 1.5)
        snack        = extend_h * 1000
        bonus        = 0  # 상여금 계산 제외 (수기 입력)

        direct_cost   = basic_a + extend_a + special_a + special_ot_a + night_amount + snack
        mgmt          = direct_cost * 0.03
        profit        = direct_cost * 0.03
        indirect_cost = mgmt + profit
        total         = math.ceil((direct_cost + indirect_cost) / 10) * 10

        c_row = 5 + emp_idx

        def _set(col_idx, value, align=RIGHT, fmt=None):
            cell = ws_c.cell(c_row, col_idx)
            cell.value          = value
            cell.alignment      = align
            cell.font           = BODY_FONT
            cell.border         = _make_border()
            if fmt:
                cell.number_format = fmt
            elif isinstance(value, (int, float)) and value and value > 100:
                cell.number_format = MONEY_FMT

        _set(2,  emp_idx+1,   CENTER)           # B: 순번
        _set(3,  name,        CENTER)           # C: 성명
        _set(4,  emp_type,    CENTER)           # D: 직원형태 (NEW)
        _set(5,  HOURLY_WAGE, RIGHT, MONEY_FMT) # E: 시급
        _set(6,  basic_h      if basic_h      > 0 else None)  # F: 기본급시간
        _set(7,  basic_a      if basic_a      > 0 else None)  # G: 기본급금액
        _set(8,  extend_h     if extend_h     > 0 else None)  # H: 연장시간
        _set(9,  extend_a     if extend_a     > 0 else None)  # I: 연장금액
        _set(10, special_h    if special_h    > 0 else None)  # J: 주특시간
        _set(11, special_a    if special_a    > 0 else None)  # K: 주특금액
        _set(12, special_ot_h if special_ot_h > 0 else None)  # L: 특잔시간
        _set(13, special_ot_a if special_ot_a > 0 else None)  # M: 특잔금액
        _set(14, night_h      if night_h      > 0 else None)  # N: 심야시간
        _set(15, night_amount if night_amount > 0 else None)  # O: 심야금액
        _set(16, snack        if snack        > 0 else None)  # P: 간식대
        _set(17, None)                                         # Q: 상여금 (빈값)
        _set(18, direct_cost  if direct_cost  > 0 else None)  # R: 직접비소계
        _set(19, int(mgmt)    if mgmt         > 0 else None)  # S: 일반관리비
        _set(20, int(profit)  if profit       > 0 else None)  # T: 기업이윤
        _set(21, int(indirect_cost) if indirect_cost > 0 else None)  # U: 간접비소계
        _set(22, total        if total        > 0 else None)  # V: 총액

    total_row = 5 + len(employees)
    sum_cell = ws_c.cell(total_row, 2, '합계')
    sum_cell.alignment = CENTER
    sum_cell.font      = BOLD9
    sum_cell.border    = _make_border()
    for blank_col in [3, 4, 5]:
        ws_c.cell(total_row, blank_col).border = _make_border()

    data_start = 5
    data_end   = 5 + len(employees) - 1

    for col_idx in [6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22]:
        col_letter = get_column_letter(col_idx)
        cell = ws_c.cell(total_row, col_idx)
        cell.value          = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        cell.alignment      = Alignment(horizontal='right', vertical='center')
        cell.font           = BOLD9
        cell.border         = _make_border()
        cell.number_format  = '#,##0'

    # C탭 금액 컬럼 천 단위 콤마
    # G(7), I(9), K(11), M(13), O(15), P(16), Q(17), R(18), S(19), T(20), U(21), V(22)
    MONEY_COLS = {7, 9, 11, 13, 15, 16, 17, 18, 19, 20, 21, 22}
    for row in ws_c.iter_rows(min_row=5, max_row=ws_c.max_row):
        for cell in row:
            if cell.column in MONEY_COLS and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT
