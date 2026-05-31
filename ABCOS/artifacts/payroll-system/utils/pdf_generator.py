import os
import math
import zipfile
import tempfile
import urllib.request
import openpyxl

from utils.errors import UserFacingError

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


FONT_NAME = 'NG'
_font_registered = False

FONT_URL = "https://github.com/google/fonts/raw/refs/heads/main/ofl/nanumgothic/NanumGothic-Regular.ttf"


def _get_font_path():
    base_dir  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    fonts_dir = os.path.join(base_dir, 'static', 'fonts')
    return os.path.join(fonts_dir, 'NanumGothic.ttf')


def _register_font():
    global _font_registered
    if _font_registered:
        return
    font_path = _get_font_path()
    if not os.path.exists(font_path):
        fonts_dir = os.path.dirname(font_path)
        os.makedirs(fonts_dir, exist_ok=True)
        try:
            urllib.request.urlretrieve(FONT_URL, font_path)
        except Exception as e:
            raise RuntimeError(
                f"나눔고딕 폰트를 다운로드하지 못했습니다: {e}\n"
                f"수동으로 NanumGothic.ttf를 {fonts_dir}에 복사해 주세요."
            )
    pdfmetrics.registerFont(TTFont(FONT_NAME, font_path))
    _font_registered = True


# ─────────────────────────────────────────────────────────
#  세금 계산 (직원형태별)
# ─────────────────────────────────────────────────────────

def _calc_withholding_tax(monthly_pay, num_dependents=1):
    """사대(A) 근로소득세 간이세액표 계산 (2026년 기준, 부양가족 수 포함)."""
    annual = monthly_pay * 12

    # 근로소득공제
    if annual <= 5_000_000:
        wd = annual * 0.70
    elif annual <= 15_000_000:
        wd = 3_500_000 + (annual - 5_000_000) * 0.40
    elif annual <= 45_000_000:
        wd = 7_500_000 + (annual - 15_000_000) * 0.15
    elif annual <= 100_000_000:
        wd = 12_000_000 + (annual - 45_000_000) * 0.05
    else:
        wd = 14_750_000 + (annual - 100_000_000) * 0.02
        wd = min(wd, 20_000_000)

    earned = max(0, annual - wd)

    # 인적공제 (부양가족 × 150만)
    personal = num_dependents * 1_500_000
    taxable  = max(0, earned - personal)

    # 누진세율 (2026)
    if taxable <= 14_000_000:
        tax = taxable * 0.06
    elif taxable <= 50_000_000:
        tax = taxable * 0.15 - 1_260_000
    elif taxable <= 88_000_000:
        tax = taxable * 0.24 - 5_760_000
    elif taxable <= 150_000_000:
        tax = taxable * 0.35 - 15_440_000
    elif taxable <= 300_000_000:
        tax = taxable * 0.38 - 19_940_000
    elif taxable <= 500_000_000:
        tax = taxable * 0.40 - 25_940_000
    else:
        tax = taxable * 0.42 - 35_940_000

    # 근로소득세액공제 (최대 74만원)
    if tax <= 1_300_000:
        credit = tax * 0.55
    else:
        credit = 715_000 + (tax - 1_300_000) * 0.30
    credit = min(credit, 740_000)

    # 표준세액공제 13만원/년
    annual_tax  = max(0, tax - credit - 130_000)
    monthly_tax = int(annual_tax / 12 / 10) * 10
    local_tax   = int(monthly_tax * 0.1 / 10) * 10

    return monthly_tax, local_tax


def _calc_tax_A(pay):
    """사대(4대보험) 2026년 기준 세금 계산."""
    national_pension = int(pay * 0.0475  / 10) * 10   # 국민연금 4.75%
    health_ins       = int(pay * 0.03595 / 10) * 10   # 건강보험 3.595%
    care_ins         = int(health_ins * 0.1314 / 10) * 10  # 장기요양 13.14%
    employ_ins       = int(pay * 0.009   / 10) * 10   # 고용보험 0.9%
    income_tax, local_tax = _calc_withholding_tax(pay)
    tax_total = national_pension + health_ins + care_ins + employ_ins + income_tax + local_tax
    return {
        'type': 'A',
        'national_pension': national_pension,
        'health_ins':       health_ins,
        'care_ins':         care_ins,
        'employ_ins':       employ_ins,
        'income_tax':       income_tax,
        'local_tax':        local_tax,
        'tax_total':        tax_total,
    }


def _calc_tax_F_high(pay):
    """프리랜서 200만원 초과: 3.3% 일괄 공제."""
    tax = int(pay * 0.033 / 10) * 10
    return {
        'type': 'F_high',
        'business_tax': tax,
        'tax_total':    tax,
    }


def _calc_tax_daily(direct_cost):
    """일용직(D) / 프리랜서 200만원 이하: 일용 공식."""
    employ_ins = int(direct_cost * 0.009 / 10) * 10

    raw_income = ((direct_cost / 7) - 150_000) * 0.027 * 7
    if raw_income < 1_000:
        income_tax = 0
    else:
        income_tax = int(raw_income / 10) * 10

    local_tax = int(income_tax * 0.1 / 10) * 10 if income_tax > 0 else 0
    tax_total = employ_ins + income_tax + local_tax
    return {
        'type': 'D',
        'employ_ins': employ_ins,
        'income_tax': income_tax,
        'local_tax':  local_tax,
        'tax_total':  tax_total,
    }


def _calc_tax(emp_type, direct_cost, total):
    """직원형태(A/D/F)에 따라 세금 계산 반환."""
    if emp_type == 'A':
        return _calc_tax_A(direct_cost)
    elif emp_type == 'D':
        return _calc_tax_daily(direct_cost)
    elif emp_type == 'F':
        if total > 2_000_000:
            return _calc_tax_F_high(total)
        else:
            return _calc_tax_daily(direct_cost)
    else:
        # 형태 미지정 → 프리랜서 기본값
        if total > 2_000_000:
            return _calc_tax_F_high(total)
        else:
            return _calc_tax_daily(direct_cost)


# ─────────────────────────────────────────────────────────
#  BC 파일 파싱
#
#  C탭 컬럼 (1-indexed, 0-indexed in row[] for parsing):
#   row[1]=B=순번  row[2]=C=성명  row[3]=D=직원형태  row[4]=E=시급
#   _v(5)=F=기본급시간  _v(6)=G=기본급금액
#   _v(7)=H=연장시간    _v(8)=I=연장금액
#   _v(9)=J=주특시간    _v(10)=K=주특금액
#   _v(11)=L=특잔시간   _v(12)=M=특잔금액
#   _v(13)=N=심야시간   _v(14)=O=심야금액
#   _v(15)=P=간식대     _v(16)=Q=상여금
#   _v(17)=R=직접비소계 _v(18)=S=일반관리비 _v(19)=T=기업이윤
#   _v(20)=U=간접비소계 _v(21)=V=총액
# ─────────────────────────────────────────────────────────

def parse_bc_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        raise UserFacingError(
            "엑셀 파일을 열 수 없습니다. 근태 및 청구내역서(xlsx) 파일인지 확인해 주세요."
        ) from exc

    # ── B탭(근태현황) 파싱 ───────────────────────────────
    b_tab = None
    for sheet_name in wb.sheetnames:
        if '근태' in sheet_name:
            b_tab = wb[sheet_name]
            break
    if b_tab is None and len(wb.sheetnames) >= 1:
        b_tab = wb[wb.sheetnames[0]]

    # B탭에서 직원별 일별 출퇴근 기록 파싱
    # 구조: DATA_ROW_0=7부터 직원당 6행씩, 컬럼 F(6)부터 31일치
    # 행 순서: 기본급(0), 연장(1), 주특(2), 특잔수당(3), 심야(4), 근태(5)
    # 이름은 B열(2), 시급은 C열(3), 직원형태는 D열(4), 구분은 E열(5)
    B_DATA_ROW_0 = 7
    B_COL_NAME   = 2
    B_COL_LABEL  = 5
    B_COL_DAY1   = 6

    b_attendance = {}  # {name: {day: {'in': h, 'ext': h, 'sp': h, 'sot': h, 'night': h, 'note': str}}}

    if b_tab is not None:
        # 직원 수 파악: 이름 컬럼이 있는 행 탐색
        row = B_DATA_ROW_0
        while True:
            name_cell = b_tab.cell(row, B_COL_NAME)
            name_val = name_cell.value
            if name_val is None or str(name_val).strip() == '':
                # 빈 이름이면 끝
                break
            name = str(name_val).strip()
            day_data = {}
            # 6개 행 파싱 (기본급/연장/주특/특잔/심야/근태)
            row_keys = ['basic', 'extend', 'special', 'special_ot', 'night', 'note']
            for r_off, key in enumerate(row_keys):
                r = row + r_off
                for d in range(1, 32):
                    col = B_COL_DAY1 + d - 1
                    cell_val = b_tab.cell(r, col).value
                    if d not in day_data:
                        day_data[d] = {k: None for k in row_keys}
                    if cell_val is not None and str(cell_val).strip() not in ('', 'nan', 'None'):
                        day_data[d][key] = cell_val
            b_attendance[name] = day_data
            row += 6

    # ── C탭(청구내역서) 파싱 ───────────────────────────────
    c_tab = None
    c_tab_name = None
    for sheet_name in wb.sheetnames:
        if '청구내역서' in sheet_name or '청구' in sheet_name:
            c_tab      = wb[sheet_name]
            c_tab_name = sheet_name
            break
    if c_tab is None:
        if len(wb.sheetnames) >= 2:
            c_tab      = wb[wb.sheetnames[1]]
            c_tab_name = wb.sheetnames[1]
        else:
            c_tab      = wb.active
            c_tab_name = wb.sheetnames[0]

    year, month = _extract_year_month(c_tab_name)

    # 헤더에서 직원형태 컬럼 위치 감지 (신규/구형 양식 자동 판별)
    new_format = _detect_new_format(c_tab)

    employees = []
    for row in c_tab.iter_rows(min_row=5):
        b_val = row[1].value if len(row) > 1 else None
        try:
            seq = int(b_val)
        except (TypeError, ValueError):
            continue

        def _v(idx, default=0):
            if len(row) > idx:
                v = row[idx].value
                if v is None:
                    return default
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return default
            return default

        name = str(row[2].value).strip() if len(row) > 2 and row[2].value else ''

        if new_format:
            # 신규 형식: D열(idx=3)=직원형태, E열(idx=4)=시급
            emp_type     = str(row[3].value).strip().upper() if len(row) > 3 and row[3].value else ''
            hourly       = _v(4, 10320)
            basic_h      = _v(5)
            basic_a      = _v(6)
            extend_h     = _v(7)
            extend_a     = _v(8)
            special_h    = _v(9)
            special_a    = _v(10)
            special_ot_h = _v(11)
            special_ot_a = _v(12)
            night_h      = _v(13)
            night_a      = _v(14)
            snack        = _v(15)
            bonus        = _v(16)
            direct_cost  = _v(17)
            total        = _v(21)
        else:
            # 구형 형식 (직원형태 컬럼 없음)
            emp_type     = ''
            hourly       = _v(3, 10320)
            basic_h      = _v(4)
            basic_a      = _v(5)
            extend_h     = _v(6)
            extend_a     = _v(7)
            special_h    = _v(8)
            special_a    = _v(9)
            special_ot_h = _v(10)
            special_ot_a = _v(11)
            night_h      = _v(12)
            night_a      = _v(13)
            snack        = _v(14)
            bonus        = _v(15)
            direct_cost  = _v(16)
            total        = _v(20)

        if int(hourly) <= 0:
            hourly = 10320
        if basic_a == 0 and basic_h > 0:
            basic_a = basic_h * hourly
        if extend_a == 0 and extend_h > 0:
            extend_a = int(extend_h * hourly * 1.5)
        if special_a == 0 and special_h > 0:
            special_a = int(special_h * hourly * 1.5)
        if special_ot_a == 0 and special_ot_h > 0:
            special_ot_a = int(special_ot_h * hourly * 1.5)
        if snack == 0 and extend_h > 0:
            snack = int(extend_h * 1000)
        if direct_cost == 0:
            direct_cost = int(basic_a + extend_a + special_a + special_ot_a + night_a + snack + bonus)
        if total == 0:
            indirect_cost = direct_cost * 0.06
            total = math.ceil((direct_cost + indirect_cost) / 10) * 10

        employees.append({
            'seq':          seq,
            'name':         name,
            'emp_type':     emp_type,
            'hourly':       int(hourly),
            'basic_h':      int(basic_h),     'basic_a':      int(basic_a),
            'extend_h':     int(extend_h),    'extend_a':     int(extend_a),
            'special_h':    int(special_h),   'special_a':    int(special_a),
            'special_ot_h': int(special_ot_h),'special_ot_a': int(special_ot_a),
            'night_h':      int(night_h),     'night_a':      int(night_a),
            'snack':        int(snack),        'bonus':        int(bonus),
            'direct_cost':  int(direct_cost),
            'total':        int(total),
            'daily':        b_attendance.get(name, {}),
        })

    if not employees:
        raise UserFacingError(
            "청구내역서에서 직원 데이터를 찾을 수 없습니다. C탭(청구내역서) 형식을 확인해 주세요."
        )

    return {'year': year, 'month': month, 'employees': employees}


def _detect_new_format(ws):
    """C탭 3~4행 헤더에서 '직원형태' 컬럼 존재 여부 확인."""
    for row in ws.iter_rows(min_row=3, max_row=4):
        for cell in row:
            if cell.value and '직원형태' in str(cell.value):
                return True
    return False


def _extract_year_month(sheet_name):
    import re, datetime
    m = re.search(r'(\d{4})[.\-/](\d{1,2})', sheet_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    now = datetime.date.today()
    return now.year, now.month


# ─────────────────────────────────────────────────────────
#  ZIP 생성 (직원형태별 폴더)
# ─────────────────────────────────────────────────────────

TYPE_FOLDER = {'A': '사대', 'D': '일용', 'F': '프리'}


def generate_pdf_zip(parsed_data, output_zip_path):
    _register_font()
    year      = parsed_data['year']
    month     = parsed_data['month']
    employees = parsed_data['employees']

    with zipfile.ZipFile(output_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for emp in employees:
            folder = TYPE_FOLDER.get(emp.get('emp_type', ''), '기타')
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp_path = tmp.name
            try:
                _generate_single_pdf(emp, year, month, tmp_path, emp.get('daily', {}))
                zf.write(tmp_path, f"{folder}/{emp['name']}_급여명세서.pdf")
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)


# ─────────────────────────────────────────────────────────
#  개인 급여명세서 PDF 생성
# ─────────────────────────────────────────────────────────

def _generate_single_pdf(emp, year, month, output_path, daily=None):
    import calendar as _calendar
    emp_type    = emp.get('emp_type', '')
    direct_cost = emp.get('direct_cost', emp['total'])
    total       = emp['total']

    if daily is None:
        daily = emp.get('daily', {})

    tax_info = _calc_tax(emp_type, direct_cost, total)
    net_pay  = total - tax_info['tax_total']

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=18*mm, leftMargin=18*mm,
        topMargin=15*mm,   bottomMargin=15*mm,
    )
    W = A4[0] - 36*mm

    BLUE   = colors.HexColor('#283593')
    LBLUE  = colors.HexColor('#E8EAF6')
    GREY   = colors.HexColor('#F5F5F5')
    WHITE  = colors.white
    GREEN  = colors.HexColor('#1B5E20')
    LGREEN = colors.HexColor('#E8F5E9')
    DKGREY = colors.HexColor('#CCCCCC')

    def fmt(n):
        if not n or n == 0:
            return '-'
        return f"{int(n):,}"

    def h(n):
        if not n or n == 0:
            return '-'
        return f"{int(n)}h"

    def p(text, size=9, bold=False, align='CENTER', color=colors.black):
        a = {'LEFT': 0, 'CENTER': 1, 'RIGHT': 2}[align]
        style = ParagraphStyle(
            's', fontName=FONT_NAME, fontSize=size,
            alignment=a, leading=size + 5, textColor=color,
            fontWeight='bold' if bold else 'normal',
        )
        return Paragraph(str(text), style)

    # ── 급여항목 테이블 ──────────────────────────────
    pay_cw = [W * 0.30, W * 0.12, W * 0.20]
    pay_data = [
        [p('급 여 항 목', 10, color=WHITE), p('시 간', 10, color=WHITE), p('금 액', 10, color=WHITE)],
        [p('기  본  급'),    p(h(emp['basic_h'])),       p(fmt(emp['basic_a']),      align='RIGHT')],
        [p('인  센  티  브'), p('-'),                     p('-')],
        [p('심  야  수  당'), p(h(emp['night_h'])),       p(fmt(emp['night_a']),      align='RIGHT')],
        [p('주      특'),    p(h(emp['special_h'])),      p(fmt(emp['special_a']),    align='RIGHT')],
        [p('특  잔  수  당'), p(h(emp['special_ot_h'])),  p(fmt(emp['special_ot_a']), align='RIGHT')],
        [p('간  식  대'),    p('-'),                      p(fmt(emp['snack']),        align='RIGHT')],
        [p('상  여  금'),    p('-'),                      p(fmt(emp['bonus']),        align='RIGHT')],
        [p('연      장'),    p(h(emp['extend_h'])),       p(fmt(emp['extend_a']),     align='RIGHT')],
        [p('급 여 총 액', 10, color=BLUE), p(''),
         p(f"{fmt(total)} 원", 10, color=BLUE, align='RIGHT')],
    ]
    pay_table = Table(pay_data, colWidths=pay_cw, repeatRows=1)
    pay_table.setStyle(TableStyle([
        ('FONTNAME',      (0, 0), (-1, -1), FONT_NAME),
        ('BACKGROUND',    (0, 0), (-1,  0), BLUE),
        ('TEXTCOLOR',     (0, 0), (-1,  0), WHITE),
        ('BACKGROUND',    (0,-1), (-1, -1), LBLUE),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',         (2, 1), (2,  -1), 'RIGHT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('LINEBELOW',     (0,-2), (-1, -2), 1.5, BLUE),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS',(0, 1), (-1, -2), [WHITE, GREY]),
    ]))

    # ── 공제항목 테이블 (직원형태별) ────────────────
    ded_cw   = [W * 0.25, W * 0.13]
    ded_rows = _build_ded_rows(tax_info, fmt, p, BLUE)
    ded_table = Table(ded_rows, colWidths=ded_cw, repeatRows=1)
    ded_table.setStyle(TableStyle([
        ('FONTNAME',      (0, 0), (-1, -1), FONT_NAME),
        ('BACKGROUND',    (0, 0), (-1,  0), BLUE),
        ('TEXTCOLOR',     (0, 0), (-1,  0), WHITE),
        ('BACKGROUND',    (0,-1), (-1, -1), LBLUE),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',         (1, 1), (1,  -1), 'RIGHT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('LINEBELOW',     (0,-2), (-1, -2), 1.5, BLUE),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS',(0, 1), (-1, -2), [WHITE, GREY]),
    ]))

    # ── 좌우 테이블 나란히 ───────────────────────────
    combined = Table(
        [[pay_table, Spacer(4*mm, 1), ded_table]],
        colWidths=[sum(pay_cw), 4*mm, sum(ded_cw)]
    )
    combined.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    # ── 실수령액 강조 박스 ───────────────────────────
    net_table = Table(
        [[p('실   수   령   액', 13, color=GREEN),
          p(f"{fmt(net_pay)} 원", 15, color=GREEN)]],
        colWidths=[W * 0.5, W * 0.5]
    )
    net_table.setStyle(TableStyle([
        ('FONTNAME',      (0, 0), (-1, -1), FONT_NAME),
        ('BACKGROUND',    (0, 0), (-1, -1), LGREEN),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX',           (0, 0), (-1, -1), 1.5, GREEN),
        ('TOPPADDING',    (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))

    # ── 안내 문구 박스 ───────────────────────────────
    footer_table = Table(
        [[p('한 달 동안 수고 많으셨습니다.^^', 9)]],
        colWidths=[W]
    )
    footer_table.setStyle(TableStyle([
        ('FONTNAME',      (0, 0), (-1, -1), FONT_NAME),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX',           (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND',    (0, 0), (-1, -1), GREY),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    # ── 일별 출퇴근 기록표 ───────────────────────────
    _, days_in_month = _calendar.monthrange(year, month)

    attend_header = [p('일', 8, color=WHITE), p('기본', 8, color=WHITE), p('연장', 8, color=WHITE), p('주특', 8, color=WHITE), p('특잔', 8, color=WHITE), p('심야', 8, color=WHITE), p('근태', 8, color=WHITE)]
    attend_rows   = [attend_header]
    DOW_KR = ['월', '화', '수', '목', '금', '토', '일']
    import datetime as _dt
    for d in range(1, days_in_month + 1):
        dow   = DOW_KR[_dt.date(year, month, d).weekday()]
        ddata = daily.get(d, {})
        def _dv(key):
            v = ddata.get(key)
            if v is None:
                return '-'
            return str(v)
        attend_rows.append([
            p(f"{d}({dow})", 7),
            p(_dv('basic'),      7),
            p(_dv('extend'),     7),
            p(_dv('special'),    7),
            p(_dv('special_ot'), 7),
            p(_dv('night'),      7),
            p(_dv('note'),       7),
        ])

    attend_cw = [W * 0.16, W * 0.12, W * 0.12, W * 0.12, W * 0.12, W * 0.12, W * 0.24]
    attend_table = Table(attend_rows, colWidths=attend_cw, repeatRows=1)
    attend_ts = [
        ('FONTNAME',      (0, 0), (-1, -1), FONT_NAME),
        ('BACKGROUND',    (0, 0), (-1,  0), BLUE),
        ('TEXTCOLOR',     (0, 0), (-1,  0), WHITE),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',          (0, 0), (-1, -1), 0.4, DKGREY),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    for i, d in enumerate(range(1, days_in_month + 1)):
        row_i = i + 1
        dow_idx = _dt.date(year, month, d).weekday()
        if dow_idx == 5:
            attend_ts.append(('BACKGROUND', (0, row_i), (-1, row_i), colors.HexColor('#DDEEFF')))
        elif dow_idx == 6:
            attend_ts.append(('BACKGROUND', (0, row_i), (-1, row_i), colors.HexColor('#FFDDDD')))
        elif i % 2 == 0:
            attend_ts.append(('BACKGROUND', (0, row_i), (-1, row_i), WHITE))
        else:
            attend_ts.append(('BACKGROUND', (0, row_i), (-1, row_i), GREY))
    attend_table.setStyle(TableStyle(attend_ts))

    company_style = ParagraphStyle('c', fontName=FONT_NAME, fontSize=11, alignment=1,
                                   spaceAfter=1*mm, textColor=colors.HexColor('#555555'))
    title_style = ParagraphStyle('t', fontName=FONT_NAME, fontSize=18, alignment=1,
                                 spaceAfter=3*mm, textColor=BLUE)
    year_style  = ParagraphStyle('y', fontName=FONT_NAME, fontSize=11, alignment=1,
                                 spaceAfter=1*mm, textColor=colors.HexColor('#555555'))
    name_style  = ParagraphStyle('n', fontName=FONT_NAME, fontSize=13, alignment=1,
                                 spaceAfter=5*mm, textColor=colors.black)
    attend_title_style = ParagraphStyle('at', fontName=FONT_NAME, fontSize=10, alignment=0,
                                        spaceBefore=4*mm, spaceAfter=2*mm, textColor=BLUE)

    elements = [
        Paragraph('어브코스', company_style),
        Paragraph('급   여   명   세   서', title_style),
        Paragraph(f'{year}년 {month}월', year_style),
        Paragraph(f"{emp['name']} 님", name_style),
        combined,
        Spacer(1, 4*mm),
        net_table,
        Spacer(1, 4*mm),
        footer_table,
        Paragraph('■ 일별 출퇴근 기록표', attend_title_style),
        attend_table,
    ]

    doc.build(elements)


def _build_ded_rows(tax_info, fmt, p, BLUE):
    """직원형태별 공제 항목 행 생성 (총 10행: 헤더1 + 데이터8 + 합계1)."""
    header = [p('공 제 항 목', 10, color=colors.white), p('금 액', 10, color=colors.white)]
    total_row = [
        p('공 제 총 액', 10, color=BLUE),
        p(f"{fmt(tax_info['tax_total'])} 원", 10, color=BLUE, align='RIGHT'),
    ]

    t = tax_info['type']

    if t == 'A':
        rows = [
            [p('국 민 연 금'),  p(fmt(tax_info['national_pension']), align='RIGHT')],
            [p('건 강 보 험'),  p(fmt(tax_info['health_ins']),       align='RIGHT')],
            [p('장기요양보험'), p(fmt(tax_info['care_ins']),         align='RIGHT')],
            [p('고  용  보  험'),p(fmt(tax_info['employ_ins']),      align='RIGHT')],
            [p('근 로 소 득 세'),p(fmt(tax_info['income_tax']),      align='RIGHT')],
            [p('지 방 소 득 세'),p(fmt(tax_info['local_tax']),       align='RIGHT')],
        ]
    elif t == 'F_high':
        rows = [
            [p('사업소득세(3.3%)'), p(fmt(tax_info['business_tax']), align='RIGHT')],
        ]
    else:  # D 또는 F 저소득
        income_str = fmt(tax_info['income_tax']) if tax_info['income_tax'] > 0 else '-'
        local_str  = fmt(tax_info['local_tax'])  if tax_info['local_tax']  > 0 else '-'
        rows = [
            [p('고  용  보  험'), p(fmt(tax_info['employ_ins']), align='RIGHT')],
            [p('소  득  세'),     p(income_str,                  align='RIGHT')],
            [p('지 방 소 득 세'), p(local_str,                   align='RIGHT')],
        ]

    # 8개 데이터 행으로 패딩
    while len(rows) < 8:
        rows.append([p(''), p('')])

    return [header] + rows + [total_row]
