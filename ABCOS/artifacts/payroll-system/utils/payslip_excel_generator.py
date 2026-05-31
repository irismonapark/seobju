import calendar
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


FONT_NAME = '맑은 고딕'

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

HEADER_FILL = PatternFill('solid', fgColor='283593')
HEADER_FONT = Font(color='FFFFFF', bold=True, name=FONT_NAME, size=9)
BOLD_FONT   = Font(name=FONT_NAME, size=9, bold=True)
NORM_FONT   = Font(name=FONT_NAME, size=9)
TOTAL_FILL  = PatternFill('solid', fgColor='E8EAF6')
NET_FILL    = PatternFill('solid', fgColor='E8F5E9')
NET_FONT    = Font(name=FONT_NAME, size=10, bold=True, color='1B5E20')

SAT_FILL = PatternFill('solid', fgColor='DDEEFF')
SUN_FILL = PatternFill('solid', fgColor='FFDDDD')
ODD_FILL = PatternFill('solid', fgColor='F5F5F5')
EVN_FILL = PatternFill('solid', fgColor='FFFFFF')

DOW_KR = ['월', '화', '수', '목', '금', '토', '일']


def _thin():
    return Side(style='thin', color='AAAAAA')


def _medium():
    return Side(style='medium', color='444444')


def _border(thick_bottom=False):
    bot = _medium() if thick_bottom else _thin()
    return Border(left=_thin(), right=_thin(), top=_thin(), bottom=bot)


def _calc_tax_for_excel(emp_type, direct_cost, total):
    """세금 계산 (pdf_generator에서 동일 로직 참조)"""
    def _withholding(monthly_pay, num_dependents=1):
        annual = monthly_pay * 12
        if annual <= 5_000_000:        wd = annual * 0.70
        elif annual <= 15_000_000:     wd = 3_500_000 + (annual - 5_000_000) * 0.40
        elif annual <= 45_000_000:     wd = 7_500_000 + (annual - 15_000_000) * 0.15
        elif annual <= 100_000_000:    wd = 12_000_000 + (annual - 45_000_000) * 0.05
        else:                          wd = min(14_750_000 + (annual - 100_000_000) * 0.02, 20_000_000)
        earned  = max(0, annual - wd)
        taxable = max(0, earned - num_dependents * 1_500_000)
        if taxable <= 14_000_000:      tax = taxable * 0.06
        elif taxable <= 50_000_000:    tax = taxable * 0.15 - 1_260_000
        elif taxable <= 88_000_000:    tax = taxable * 0.24 - 5_760_000
        elif taxable <= 150_000_000:   tax = taxable * 0.35 - 15_440_000
        elif taxable <= 300_000_000:   tax = taxable * 0.38 - 19_940_000
        elif taxable <= 500_000_000:   tax = taxable * 0.40 - 25_940_000
        else:                          tax = taxable * 0.42 - 35_940_000
        credit = min(715_000 + (tax - 1_300_000) * 0.30 if tax > 1_300_000 else tax * 0.55, 740_000)
        annual_tax = max(0, tax - credit - 130_000)
        monthly_tax = int(annual_tax / 12 / 10) * 10
        local_tax   = int(monthly_tax * 0.1 / 10) * 10
        return monthly_tax, local_tax

    if emp_type == 'A':
        np_  = int(direct_cost * 0.0475  / 10) * 10
        hi   = int(direct_cost * 0.03595 / 10) * 10
        ci   = int(hi * 0.1314 / 10) * 10
        ei   = int(direct_cost * 0.009   / 10) * 10
        it_, lt_ = _withholding(direct_cost)
        return np_ + hi + ci + ei + it_ + lt_
    elif emp_type == 'D':
        ei = int(direct_cost * 0.009 / 10) * 10
        raw = ((direct_cost / 7) - 150_000) * 0.027 * 7
        it_ = int(raw / 10) * 10 if raw >= 1_000 else 0
        lt_ = int(it_ * 0.1 / 10) * 10 if it_ > 0 else 0
        return ei + it_ + lt_
    elif emp_type == 'F':
        if total > 2_000_000:
            return int(total * 0.033 / 10) * 10
        else:
            ei = int(direct_cost * 0.009 / 10) * 10
            raw = ((direct_cost / 7) - 150_000) * 0.027 * 7
            it_ = int(raw / 10) * 10 if raw >= 1_000 else 0
            lt_ = int(it_ * 0.1 / 10) * 10 if it_ > 0 else 0
            return ei + it_ + lt_
    else:
        if total > 2_000_000:
            return int(total * 0.033 / 10) * 10
        else:
            ei = int(direct_cost * 0.009 / 10) * 10
            raw = ((direct_cost / 7) - 150_000) * 0.027 * 7
            it_ = int(raw / 10) * 10 if raw >= 1_000 else 0
            lt_ = int(it_ * 0.1 / 10) * 10 if it_ > 0 else 0
            return ei + it_ + lt_


def _set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    cell = ws.cell(row, col, value)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell


def _build_cover_sheet(ws, employees, year, month, company=''):
    """겉지 탭 생성: 순번·이름·금액·급여대장·확인·비고"""
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    company_suffix = f'_{company}' if company else ''
    title_cell = ws.cell(1, 1, f"{year}년 {month:02d}월 급여명세서{company_suffix}")
    title_cell.font      = Font(name=FONT_NAME, size=13, bold=True, color='283593')
    title_cell.alignment = CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 20

    headers = ['순번', '이름', '금액', '급여대장', '확인', '비고']
    for ci, h in enumerate(headers, 1):
        _set_cell(ws, 2, ci, h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER, border=_border())
    ws.row_dimensions[2].height = 16

    total_amount = 0
    for idx, emp in enumerate(employees, 1):
        r     = 2 + idx
        total = emp['total']
        total_amount += total
        _set_cell(ws, r, 1, idx,          font=NORM_FONT, align=CENTER, border=_border())
        _set_cell(ws, r, 2, emp['name'],  font=NORM_FONT, align=CENTER, border=_border())
        _set_cell(ws, r, 3, total,        font=NORM_FONT, align=RIGHT,  border=_border(), num_fmt='#,##0')
        _set_cell(ws, r, 4, None,         font=NORM_FONT, align=CENTER, border=_border())
        _set_cell(ws, r, 5, None,         font=NORM_FONT, align=CENTER, border=_border())
        _set_cell(ws, r, 6, None,         font=NORM_FONT, align=CENTER, border=_border())
        ws.row_dimensions[r].height = 15

    total_row = 2 + len(employees) + 1
    _set_cell(ws, total_row, 1, '합계',       font=BOLD_FONT,  align=CENTER, border=_border(True))
    _set_cell(ws, total_row, 2, None,         font=BOLD_FONT,  align=CENTER, border=_border(True))
    _set_cell(ws, total_row, 3, total_amount, font=BOLD_FONT,  align=RIGHT,  border=_border(True), num_fmt='#,##0')
    _set_cell(ws, total_row, 4, None,         font=NORM_FONT,  align=CENTER, border=_border(True))
    _set_cell(ws, total_row, 5, None,         font=NORM_FONT,  align=CENTER, border=_border(True))
    _set_cell(ws, total_row, 6, None,         font=NORM_FONT,  align=CENTER, border=_border(True))
    ws.row_dimensions[total_row].height = 16


def _build_emp_sheet(ws, emp, year, month, company=''):
    """직원별 탭: 2단 급여명세 + 가로 일자별 근태 그리드 (이미지 레이아웃)"""
    _, days_in_month = calendar.monthrange(year, month)

    # ── 컬럼 너비 설정 ─────────────────────────────────
    # A(1): 이름세로 / 급여항목 레이블
    # B(2): 구분 / 시간
    # C(3)~AG(33): 일자 1~31 (narrow)
    # AH(34): 소급
    # AI(35): 합계
    # 상단 급여테이블은 A-AI 컬럼을 merge하여 사용
    COL_NAME_V = 1   # A: 이름(세로) / 상단 급여항목 레이블
    COL_LABEL  = 2   # B: 구분 / 상단 시간
    COL_DAY1   = 3   # C: 1일
    COL_SOGEUP = COL_DAY1 + 31   # 34
    COL_SUM    = COL_DAY1 + 32   # 35

    ws.column_dimensions['A'].width = 5    # 이름세로 / 급여항목
    ws.column_dimensions['B'].width = 9    # 구분 / 시간
    # 일자 컬럼 C(3) ~ AG(33)
    for i in range(31):
        ws.column_dimensions[get_column_letter(COL_DAY1 + i)].width = 3.5
    # 소급, 합계
    ws.column_dimensions[get_column_letter(COL_SOGEUP)].width = 5
    ws.column_dimensions[get_column_letter(COL_SUM)].width = 8

    name     = emp['name']
    emp_type = emp.get('emp_type', '')
    total    = emp['total']
    dc       = emp.get('direct_cost', total)
    tax      = _calc_tax_for_excel(emp_type, dc, total)
    net_pay  = total - tax

    company_suffix = f'_{company}' if company else ''

    # 상단 섹션에서 쓸 컬럼 range
    # 좌측 급여: A-P(1-16) → 16 컬럼을 3단으로 나눔
    # 우측 공제: Q-AI(17-35) → 19 컬럼을 2단으로 나눔
    # 각 셀은 merge로 너비 확보

    # 급여 좌측: 항목=A..F(1-6 merge), 시간=G..J(7-10 merge), 금액=K..P(11-16 merge)
    # 세액 우측: 항목=Q..W(17-23 merge), 금액=X..AI(24-35 merge)
    PAY_LABEL_START  = 1
    PAY_LABEL_END    = 7
    PAY_TIME_START   = 8
    PAY_TIME_END     = 11
    PAY_AMT_START    = 12
    PAY_AMT_END      = 18
    DED_LABEL_START  = 19
    DED_LABEL_END    = 27
    DED_AMT_START    = 28
    DED_AMT_END      = COL_SUM   # 35

    # Total columns end = COL_SUM = 35

    def _merge_hdr(row, c1, c2, value, fill=HEADER_FILL, font=HEADER_FONT, align=CENTER):
        for c in range(c1, c2 + 1):
            ws.cell(row, c).fill   = fill
            ws.cell(row, c).border = _border()
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row, c1, value)
        cell.font = font; cell.fill = fill; cell.alignment = align; cell.border = _border()
        return cell

    def _merge_data(row, c1, c2, value, fill, font=NORM_FONT, align=CENTER, num_fmt=None):
        for c in range(c1, c2 + 1):
            ws.cell(row, c).fill   = fill
            ws.cell(row, c).border = _border()
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row, c1, value)
        cell.font = font; cell.fill = fill; cell.alignment = align; cell.border = _border()
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    # ── Row 1: 제목 ──────────────────────────────────────
    title_val = f"급여명세서  {year}년 {month:02d}월{company_suffix}"
    _merge_hdr(1, 1, COL_SUM, title_val,
               fill=PatternFill('solid', fgColor='FFFFFF'),
               font=Font(name=FONT_NAME, size=12, bold=True, color='283593'),
               align=CENTER)
    ws.row_dimensions[1].height = 20

    # ── Row 2: 섹션 헤더 ──────────────────────────────────
    _merge_hdr(2, PAY_LABEL_START, PAY_AMT_END, '급  여')
    _merge_hdr(2, DED_LABEL_START, DED_AMT_END, '세  액  및  공  제')
    ws.row_dimensions[2].height = 14

    # ── Row 3: 열 헤더 ────────────────────────────────────
    _merge_hdr(3, PAY_LABEL_START, PAY_LABEL_END, '항  목')
    _merge_hdr(3, PAY_TIME_START,  PAY_TIME_END,  '시  간')
    _merge_hdr(3, PAY_AMT_START,   PAY_AMT_END,   '금  액')
    _merge_hdr(3, DED_LABEL_START, DED_LABEL_END, '항  목')
    _merge_hdr(3, DED_AMT_START,   DED_AMT_END,   '금  액')
    ws.row_dimensions[3].height = 14

    # ── 급여 항목 (좌측) ─────────────────────────────────
    pay_items = [
        ('기 본 급',      emp.get('basic_h'),       emp.get('basic_a')),
        ('잔 업 수 당',   emp.get('extend_h'),      emp.get('extend_a')),
        ('휴 일 수 당',   emp.get('special_h'),     emp.get('special_a')),
        ('특 근 수 당',   emp.get('special_ot_h'),  emp.get('special_ot_a')),
        ('심 야 수 당',   emp.get('night_h'),       emp.get('night_a')),
        ('지 각 / 조 퇴', None,                    None),
        ('교 대 주 휴',   None,                     None),
        ('만 근 수 당',   None,                     None),
        ('상 여 금',      None,                     emp.get('bonus')),
        ('간 식 대',      None,                     emp.get('snack')),
    ]

    # ── 세액및공제 항목 (우측) ───────────────────────────
    ded_items = [
        ('갑 근 세',    None),
        ('소 득 세',    None),
        ('건 강 보 험', None),
        ('장 기 요 양', None),
        ('국 민 연 금', None),
        ('공 과 금',    None),
        ('가 불 금',    None),
        ('선 지 급',    None),
    ]

    max_rows = max(len(pay_items), len(ded_items))

    for i in range(max_rows):
        r    = 4 + i
        fill = ODD_FILL if i % 2 == 0 else EVN_FILL

        if i < len(pay_items):
            p_label, p_hrs, p_amt = pay_items[i]
            _merge_data(r, PAY_LABEL_START, PAY_LABEL_END, p_label, fill, align=CENTER)
            _merge_data(r, PAY_TIME_START,  PAY_TIME_END,
                        int(p_hrs) if p_hrs and p_hrs > 0 else None, fill, align=CENTER)
            _merge_data(r, PAY_AMT_START,   PAY_AMT_END,
                        int(p_amt) if p_amt and p_amt > 0 else None, fill,
                        align=RIGHT, num_fmt='#,##0')
        else:
            _merge_data(r, PAY_LABEL_START, PAY_AMT_END, None, fill)

        if i < len(ded_items):
            d_label, d_amt = ded_items[i]
            _merge_data(r, DED_LABEL_START, DED_LABEL_END, d_label, fill, align=CENTER)
            _merge_data(r, DED_AMT_START,   DED_AMT_END,
                        int(d_amt) if d_amt and d_amt > 0 else None, fill,
                        align=RIGHT, num_fmt='#,##0')
        else:
            _merge_data(r, DED_LABEL_START, DED_AMT_END, None, fill)

        ws.row_dimensions[r].height = 14

    # ── 합계 행 ───────────────────────────────────────────
    sum_row = 4 + max_rows
    _merge_data(sum_row, PAY_LABEL_START, PAY_LABEL_END, '급 여 총 액', TOTAL_FILL, font=BOLD_FONT)
    _merge_data(sum_row, PAY_TIME_START,  PAY_TIME_END,  None,         TOTAL_FILL, font=BOLD_FONT)
    _merge_data(sum_row, PAY_AMT_START,   PAY_AMT_END,   total,        TOTAL_FILL,
                font=BOLD_FONT, align=RIGHT, num_fmt='#,##0')
    _merge_data(sum_row, DED_LABEL_START, DED_LABEL_END, '공 제 총 액', TOTAL_FILL, font=BOLD_FONT)
    _merge_data(sum_row, DED_AMT_START,   DED_AMT_END,   tax,          TOTAL_FILL,
                font=BOLD_FONT, align=RIGHT, num_fmt='#,##0')
    ws.row_dimensions[sum_row].height = 16

    net_row = sum_row + 1
    _merge_data(net_row, PAY_LABEL_START, PAY_LABEL_END, '실 수 령 액', NET_FILL, font=NET_FONT)
    _merge_data(net_row, PAY_TIME_START,  PAY_TIME_END,  None,          NET_FILL, font=NET_FONT)
    _merge_data(net_row, PAY_AMT_START,   PAY_AMT_END,   net_pay,       NET_FILL,
                font=NET_FONT, align=RIGHT, num_fmt='#,##0')
    _merge_data(net_row, DED_LABEL_START, DED_LABEL_END, '실 수 령 액', NET_FILL, font=NET_FONT)
    _merge_data(net_row, DED_AMT_START,   DED_AMT_END,   net_pay,       NET_FILL,
                font=NET_FONT, align=RIGHT, num_fmt='#,##0')
    ws.row_dimensions[net_row].height = 18

    # ── 감사 문구 행 ──────────────────────────────────────
    msg_row = net_row + 1
    PEACH_FILL = PatternFill('solid', fgColor='FFE4C4')
    _merge_hdr(msg_row, 1, COL_SUM, '한 달 동안 수고 많으셨습니다 ^^',
               fill=PEACH_FILL,
               font=Font(name=FONT_NAME, size=10, bold=True, color='8B0000'),
               align=CENTER)
    ws.row_dimensions[msg_row].height = 16

    # ── 이름 행 ───────────────────────────────────────────
    name_row = msg_row + 1
    _merge_data(name_row, 1, 3, '이름 :',
                PatternFill('solid', fgColor='FFFFFF'),
                font=Font(name=FONT_NAME, size=10, bold=True), align=RIGHT)
    _merge_data(name_row, 4, COL_SUM, name,
                PatternFill('solid', fgColor='FFFFFF'),
                font=Font(name=FONT_NAME, size=12, bold=True, color='1A237E'), align=CENTER)
    ws.row_dimensions[name_row].height = 20

    # ── 빈 행 구분 ────────────────────────────────────────
    gap_row = name_row + 1
    ws.row_dimensions[gap_row].height = 6

    # ── 가로 일자별 근태 그리드 ──────────────────────────
    daily = emp.get('daily', {})

    grid_hdr_row = gap_row + 1    # 일자 헤더
    grid_dow_row = grid_hdr_row + 1  # 요일 헤더
    grid_data_row = grid_dow_row + 1  # 기본부터

    row_labels = ['기본', '연장', '주특', '특잔', '심야', '근태']
    row_keys   = ['basic', 'extend', 'special', 'special_ot', 'night', 'attendance']

    total_grid_rows = len(row_labels)

    # 이름 컬럼 (A): merge 전체 행
    ws.merge_cells(
        start_row=grid_hdr_row, start_column=COL_NAME_V,
        end_row=grid_data_row + total_grid_rows - 1, end_column=COL_NAME_V
    )
    name_v_cell = ws.cell(grid_hdr_row, COL_NAME_V, name)
    name_v_cell.font      = Font(name=FONT_NAME, size=9, bold=True)
    name_v_cell.alignment = Alignment(horizontal='center', vertical='center',
                                      textRotation=90, wrap_text=True)
    name_v_cell.border    = _border()
    name_v_cell.fill      = HEADER_FILL
    name_v_cell.font      = HEADER_FONT

    # 구분/일자 헤더 행
    _set_cell(ws, grid_hdr_row, COL_LABEL, '구분',
              font=HEADER_FONT, fill=HEADER_FILL, align=CENTER, border=_border())
    ws.merge_cells(start_row=grid_hdr_row, start_column=COL_LABEL,
                   end_row=grid_dow_row, end_column=COL_LABEL)

    # 일자 번호 헤더
    for d in range(1, days_in_month + 1):
        col = COL_DAY1 + d - 1
        dow_i = datetime.date(year, month, d).weekday()
        dtype = 'saturday' if dow_i == 5 else ('sunday' if dow_i == 6 else 'weekday')
        if dtype == 'saturday':
            hdr_fill = SAT_FILL
            hdr_font = Font(color='0000CC', name=FONT_NAME, size=8, bold=True)
        elif dtype == 'sunday':
            hdr_fill = SUN_FILL
            hdr_font = Font(color='CC0000', name=FONT_NAME, size=8, bold=True)
        else:
            hdr_fill = HEADER_FILL
            hdr_font = Font(color='FFFFFF', name=FONT_NAME, size=8, bold=True)
        c = ws.cell(grid_hdr_row, col, d)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = CENTER; c.border = _border()

    # 빈 날짜 채우기 (days_in_month+1 ~ 31)
    for d in range(days_in_month + 1, 32):
        col = COL_DAY1 + d - 1
        ws.cell(grid_hdr_row, col).border = _border()

    _set_cell(ws, grid_hdr_row, COL_SOGEUP, '소급',
              font=HEADER_FONT, fill=HEADER_FILL, align=CENTER, border=_border())
    _set_cell(ws, grid_hdr_row, COL_SUM, '합계',
              font=HEADER_FONT, fill=HEADER_FILL, align=CENTER, border=_border())

    # 요일 헤더 행
    ws.row_dimensions[grid_hdr_row].height = 14
    ws.row_dimensions[grid_dow_row].height = 13
    for d in range(1, days_in_month + 1):
        col   = COL_DAY1 + d - 1
        dow_i = datetime.date(year, month, d).weekday()
        dow   = DOW_KR[dow_i]
        if dow_i == 5:
            df = SAT_FILL; font = Font(color='0000CC', name=FONT_NAME, size=8)
        elif dow_i == 6:
            df = SUN_FILL; font = Font(color='CC0000', name=FONT_NAME, size=8)
        else:
            df = HEADER_FILL; font = Font(color='FFFFFF', name=FONT_NAME, size=8)
        c = ws.cell(grid_dow_row, col, dow)
        c.font = font; c.fill = df
        c.alignment = CENTER; c.border = _border()
    for d in range(days_in_month + 1, 32):
        ws.cell(grid_dow_row, COL_DAY1 + d - 1).border = _border()
    ws.cell(grid_dow_row, COL_SOGEUP).border = _border()
    ws.cell(grid_dow_row, COL_SUM).border = _border()

    # 데이터 행
    for r_off, (label, key) in enumerate(zip(row_labels, row_keys)):
        r    = grid_data_row + r_off
        fill = ODD_FILL if r_off % 2 == 0 else EVN_FILL
        ws.row_dimensions[r].height = 14

        _set_cell(ws, r, COL_LABEL, label,
                  font=BOLD_FONT, fill=HEADER_FILL if key == 'attendance' else fill,
                  align=CENTER, border=_border())

        row_total = 0
        for d in range(1, 32):
            col  = COL_DAY1 + d - 1
            cell = ws.cell(r, col)
            cell.alignment = CENTER
            cell.border    = _border()
            cell.font      = NORM_FONT

            if d <= days_in_month:
                dow_i = datetime.date(year, month, d).weekday()
                if dow_i == 5:
                    cell.fill = SAT_FILL
                elif dow_i == 6:
                    cell.fill = SUN_FILL
                else:
                    cell.fill = fill

                ddata = daily.get(d, {})
                if key == 'attendance':
                    att_val = ddata.get('note')
                    cell.value = att_val
                    cell.fill  = PatternFill('solid', fgColor='FFFF99') if att_val else (
                        SAT_FILL if dow_i == 5 else (SUN_FILL if dow_i == 6 else fill))
                elif key == 'night':
                    v = ddata.get('night', 0) or 0
                    cell.value = int(v) if v and v > 0 else None
                    if cell.value:
                        row_total += int(v)
                else:
                    v = ddata.get(key, 0) or 0
                    cell.value = int(v) if v > 0 else None
                    if cell.value:
                        row_total += int(v)
            else:
                cell.fill = PatternFill('solid', fgColor='DDDDDD')

        ws.cell(r, COL_SOGEUP).border = _border()
        ws.cell(r, COL_SOGEUP).fill  = fill

        sum_cell = ws.cell(r, COL_SUM)
        sum_cell.border = _border()
        sum_cell.fill   = TOTAL_FILL
        sum_cell.font   = BOLD_FONT
        sum_cell.alignment = CENTER
        if key == 'attendance':
            sum_cell.value = '계'
        else:
            sum_cell.value = row_total if row_total > 0 else None


EMP_TYPE_LABEL = {'A': '사대', 'F': '프리', 'D': '일용'}


def generate_payslip_excel_zip(parsed_data, output_zip_path, company=''):
    """BC 파일에서 파싱한 데이터로 사대/프리/일용 각각의 엑셀 파일을 ZIP으로 생성."""
    import zipfile, tempfile, os

    year      = parsed_data['year']
    month     = parsed_data['month']
    employees = parsed_data['employees']

    groups = {'A': [], 'F': [], 'D': []}
    for emp in employees:
        t = emp.get('emp_type', '').upper()
        if t in groups:
            groups[t].append(emp)
        else:
            groups['A'].append(emp)

    company_suffix = f'_{company}' if company else ''

    with zipfile.ZipFile(output_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for type_key, label in EMP_TYPE_LABEL.items():
            emps = groups[type_key]

            wb = openpyxl.Workbook()

            cover_ws = wb.active
            cover_ws.title = '겉지'
            _build_cover_sheet(cover_ws, emps, year, month, company=company)

            for emp in emps:
                sheet_name = emp['name'][:31]
                ws = wb.create_sheet(title=sheet_name)
                _build_emp_sheet(ws, emp, year, month, company=company)

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
            try:
                wb.save(tmp_path)
                zf.write(tmp_path, f"{year}-{month:02d} {label} 급여명세서{company_suffix}.xlsx")
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
