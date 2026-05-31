"""A4 출력용 엑셀 레이아웃: 병합 유지, 열 너비, 테두리, 인쇄 설정"""
from datetime import date

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from template_config import EXTRA_CELL, INTRO_CELL, LAST_LAYOUT_ROW, PRINT_AREA, SIGNATURE_CELL

THIN = Side(style='thin', color='000000')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor='D9E2F3')
TITLE_FONT = Font(name='맑은 고딕', size=18, bold=True)
HEADER_FONT = Font(name='맑은 고딕', size=10, bold=True)
SIDE_LABEL_FONT = Font(name='맑은 고딕', size=10, bold=True)
BODY_FONT = Font(name='맑은 고딕', size=10)

CENTER_HEADER_CELLS = (
    'B8', 'E8', 'K8', 'O8', 'S8',
    'B12', 'E12', 'J12', 'S12',
    'B19', 'H19', 'K19',
)

EXTRA_TITLE_CELL = 'A23'
LICENSE_LABEL_CELL = 'A19'

COLUMN_WIDTHS = {
    'A': 5.5,
    'B': 12,
    'C': 2.5,
    'D': 2.5,
    'E': 13,
    'F': 2.5,
    'G': 2.5,
    'H': 2.5,
    'I': 3,
    'J': 13,
    'K': 10,
    'L': 2.5,
    'M': 2.5,
    'N': 2.5,
    'O': 7,
    'P': 2.5,
    'Q': 2.5,
    'R': 2.5,
    'S': 9,
    'T': 2.5,
}

HEADER_ROWS = (8, 12, 19)
WRAP_TOP_CELLS = (INTRO_CELL, EXTRA_CELL)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
SIDE_LABEL_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)


def resolve_write_coordinate(ws, cell_coordinate):
    """병합 셀이면 좌상단(기준) 셀 좌표 반환 — 병합 해제 없음"""
    from openpyxl.cell.cell import MergedCell

    cell = ws[cell_coordinate]
    if not isinstance(cell, MergedCell):
        return cell_coordinate

    for merged in ws.merged_cells.ranges:
        if cell_coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col).coordinate
    return cell_coordinate


def write_cell(ws, cell_coordinate, value):
    anchor = resolve_write_coordinate(ws, cell_coordinate)
    ws[anchor] = value


def clear_cell(ws, cell_coordinate):
    write_cell(ws, cell_coordinate, None)


def split_education_line(text):
    """예: '대학(2,3년제) 졸업' → (학교명, 졸업구분)"""
    line = (text or '').strip()
    if not line:
        return '', ''
    for grad in ('졸업', '재학', '수료', '중퇴', '휴학'):
        if line.endswith(grad):
            school = line[: -len(grad)].strip()
            return school or line, grad
        marker = f' {grad}'
        if marker in line:
            school = line.split(marker, 1)[0].strip()
            return school or line, grad
    return line, ''


def format_signature_line(name):
    today = date.today()
    return (
        f'{today.year}년  {today.month}월  {today.day}일    '
        f'지 원 자 : {name}  (인)'
    )


def _unmerge_if_exists(ws, range_string):
    for merged in list(ws.merged_cells.ranges):
        if str(merged) == range_string:
            ws.unmerge_cells(range_string)
            return True
    return False


def _shrink_license_section(ws):
    """면허자격 세로 라벨 1줄 축소 (A19:A21 → A19:A20)"""
    _unmerge_if_exists(ws, 'A19:A21')
    if 'A19:A20' not in [str(m) for m in ws.merged_cells.ranges]:
        ws.merge_cells('A19:A20')

    cell = ws[LICENSE_LABEL_CELL]
    cell.value = '면허\n자격'
    cell.fill = HEADER_FILL
    cell.font = SIDE_LABEL_FONT
    cell.alignment = SIDE_LABEL_ALIGN

    ws.row_dimensions[21].height = 18
    ws.row_dimensions[22].height = 6


def _style_extra_title(ws):
    """특이사항 타이틀 — 면허자격과 동일 배경·볼드, 2줄 표시"""
    cell = ws[EXTRA_TITLE_CELL]
    cell.value = '특이\n사항'
    cell.fill = HEADER_FILL
    cell.font = SIDE_LABEL_FONT
    cell.alignment = SIDE_LABEL_ALIGN

    for row in (23, 24, 25):
        ws.row_dimensions[row].height = 28


def _style_signature_area(ws):
    anchor = resolve_write_coordinate(ws, SIGNATURE_CELL)
    ws[anchor].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in (27, 28):
        ws.row_dimensions[row].height = 24


def apply_border_to_range(ws, min_row, max_row, min_col, max_col):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = THIN_BORDER


def apply_workbook_layout(ws):
    """양식 병합·인쇄 영역은 유지하고 서식·A4 정리"""
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    for col_idx in range(21, 40):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions[letter]
        dim.width = 2
        dim.hidden = True

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8
    for row in range(3, LAST_LAYOUT_ROW + 1):
        if ws.row_dimensions[row].height is None or ws.row_dimensions[row].height < 18:
            ws.row_dimensions[row].height = 22

    _shrink_license_section(ws)
    _style_extra_title(ws)
    _style_signature_area(ws)

    ws.print_area = PRINT_AREA
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6,
        header=0.3, footer=0.3,
    )

    for merged in ws.merged_cells.ranges:
        if merged.max_row > LAST_LAYOUT_ROW + 5:
            continue
        apply_border_to_range(
            ws, merged.min_row, merged.max_row, merged.min_col, merged.max_col,
        )

    for col in range(1, 21):
        for row in (1, 2):
            cell = ws.cell(row, col)
            cell.font = TITLE_FONT
            cell.alignment = CENTER_ALIGN

    for row in HEADER_ROWS:
        for col in range(1, 21):
            cell = ws.cell(row, col)
            if cell.value:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_ALIGN

    for coord in CENTER_HEADER_CELLS:
        cell = ws[coord]
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN

    for row in range(3, 7):
        for col in range(5, 21):
            cell = ws.cell(row, col)
            if cell.value and row == 3 and col >= 9:
                cell.font = BODY_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    for row in range(9, 22):
        for col in range(2, 21):
            cell = ws.cell(row, col)
            if row not in HEADER_ROWS:
                cell.font = BODY_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    for coord in WRAP_TOP_CELLS:
        anchor = resolve_write_coordinate(ws, coord)
        ws[anchor].alignment = Alignment(
            horizontal='left', vertical='top', wrap_text=True,
        )

    ws['A1'].font = TITLE_FONT
