"""sample.xlsx → resume_template.xlsx (쓰기 가능한 빈 양식)"""
import openpyxl
from openpyxl import Workbook
from pathlib import Path

BASE = Path(__file__).parent
SOURCE = BASE / 'templates' / 'sample.xlsx'
TARGET = BASE / 'templates' / 'resume_template.xlsx'

CLEAR_CELLS = [
    'I3', 'I4', 'I5', 'I6',
    'B9', 'E9', 'K9', 'B10', 'E10', 'K10', 'B11', 'E11', 'K11',
    'B13', 'E13', 'J13', 'B14', 'E14', 'J14', 'B15', 'E15', 'J15',
    'B16', 'E16', 'J16', 'B17', 'E17', 'J17', 'B18', 'E18', 'J18',
    'B20', 'E20', 'H20', 'K20', 'B21', 'E21', 'K21',
    'B23', 'K23', 'N23', 'K24', 'N24', 'A26', 'A27',
]


def main():
    src = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    src_ws = src.active

    wb = Workbook()
    ws = wb.active
    ws.title = src_ws.title

    for row in src_ws.iter_rows(min_row=1, max_row=41, max_col=14):
        for cell in row:
            if cell.value is not None:
                ws.cell(row=cell.row, column=cell.column, value=cell.value)

    src.close()

    for addr in CLEAR_CELLS:
        ws[addr] = None

    wb.save(TARGET)
    print(f'Created {TARGET}')


if __name__ == '__main__':
    main()
