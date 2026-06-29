"""기능 검증: python test_app.py"""
import os

import openpyxl

from albamon_parser import parse_albamon_resume
from app import (
    app,
    extract_text_from_pdf,
    fill_excel_template,
    make_output_filename,
    parse_resume_data,
    validate_parsed_data,
)

SAMPLE_TEXT = (
    "이름: 홍길동\n"
    "연락처: 010-1234-5678\n"
    "이메일: hong@example.com\n"
    "생년월일: 1990년 5월 15일\n"
    "주소: 서울시 강남구 테헤란로 123\n"
    "\n"
    "학력\n"
    "서울대학교 경영학과 졸업 (2015)\n"
    "\n"
    "경력\n"
    "(주)테스트컴퍼니 마케팅팀 (2016.03 ~ 2020.12)\n"
    "\n"
    "자격증\n"
    "컴활 1급\n"
)

ALBAMON_PDF = r'c:\Users\paulc\Downloads\ResumeConverter\알바몬_sample.pdf'


def test_make_output_filename():
    assert make_output_filename("홍길동") == "알바몬_홍길동님_이력서.xlsx"
    assert make_output_filename("김슬아") == "알바몬_김슬아님_이력서.xlsx"


def test_parse_and_fill():
    data = parse_resume_data(SAMPLE_TEXT)
    validate_parsed_data(data)
    out_path, out_name = fill_excel_template(data)
    assert out_name == "알바몬_홍길동님_이력서.xlsx"
    assert os.path.exists(out_path)
    wb = openpyxl.load_workbook(out_path)
    assert wb.active["I3"].value == "홍길동"
    os.remove(out_path)


def test_albamon_kimseula():
    if not os.path.isfile(ALBAMON_PDF):
        print("SKIP albamon PDF (file not found)")
        return

    text = extract_text_from_pdf(ALBAMON_PDF)
    data = parse_albamon_resume(text)
    assert data["이름"] == "김슬아"
    assert data["연락처"] == "010-6232-0450"
    assert data["생년월일"] == "1999년생"
    assert "김포" in data["주소"]
    assert len(data["경력"]) == 2
    assert data["경력"][0]["company"] == "현대카드"
    assert data["경력"][0]["duty"] == "인콜업무 및 총무업무"
    assert data["경력"][1]["company"] == "아디다스 코리아"

    out_path, out_name = fill_excel_template(data)
    assert out_name == "알바몬_김슬아님_이력서.xlsx"
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws["I3"].value == "김슬아"
    assert ws["I4"].value == "1999년생"
    assert ws["I6"].value == "010-6232-0450"
    assert ws["B13"].value == "2022.02 ~ 2023.02"
    assert ws["E13"].value == "현대카드"
    assert ws["J13"].value == "인콜업무 및 총무업무"
    assert ws["E14"].value == "아디다스 코리아"
    assert ws["J14"].value == "매장관리"
    assert ws["E9"].value == "대학(2,3년제)"
    assert ws["O9"].value == "졸업"
    assert ws["B23"].value == "항상 모든 일에 최선을 다하겠습니다!"
    assert ws["A23"].value == "자기\n소개"
    assert ws["A26"].value in (None, "")
    assert "나만의 스킬" not in (ws["B23"].value or "")
    assert "어학능력" not in (ws["B23"].value or "")
    assert len(list(ws.merged_cells.ranges)) >= 65
    os.remove(out_path)


def test_convert_requires_pdf():
    client = app.test_client()
    response = client.post("/convert", data={}, content_type="multipart/form-data")
    assert response.status_code == 400


def test_index_page():
    client = app.test_client()
    response = client.get("/")
    html = response.get_data(as_text=True)
    assert "엑셀파일로 변환" in html
    assert "template_excel" not in html


if __name__ == "__main__":
    test_make_output_filename()
    test_parse_and_fill()
    test_albamon_kimseula()
    test_convert_requires_pdf()
    test_index_page()
    print("ALL_TESTS_PASSED")
